#!/usr/bin/env python3
"""Download Scout schedules and class requirements from scoutingevent.com QR links.

The site currently returns the data successfully but its old jQuery Mobile renderer can
fail before the schedule list is initialized. This CLI uses a real browser only to
establish the PHP session and trigger the schedule request, then reads the JSON payloads
directly.
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import html
import json
import re
import sys
import time
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import parse_qs

from openpyxl import load_workbook
from playwright.async_api import Browser, BrowserContext, Page, Response, async_playwright

AJAX_URL = "https://scoutingevent.com/mobile/inc/ajax_mobile.php"
DEAD_ASSET_HOSTS = {
    "mike.dev.admin.247scouting.com",
    "mike.dev.scoutingevent.com",
}


@dataclass(slots=True)
class ScoutInput:
    name: str
    url: str
    registrant_type: str = ""


@dataclass(slots=True)
class ClassRecord:
    scout_name: str
    attendee_id: str
    class_number: str
    class_p4_id: str
    class_name: str
    days: str
    period_name: str
    start_time: str
    start_time_24h: str
    location: str
    has_requirements: bool
    class_completed: bool
    raw_class_name: str


@dataclass(slots=True)
class RequirementRecord:
    scout_name: str
    attendee_id: str
    class_number: str
    class_p4_id: str
    merit_badge_name: str
    requirement_id: str
    requirement_type_id: str
    requirement_number: str
    description: str
    completed_requirement: bool
    completed_class: bool
    requirement_on_event: bool
    present_days: str


@dataclass(slots=True)
class ScoutResult:
    name: str
    qr_url: str
    attendee_id: str = ""
    classes: list[ClassRecord] = field(default_factory=list)
    requirements: list[RequirementRecord] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Pull schedules and class requirements from ScoutingEvent QR URLs, "
            "bypassing the site's broken jQuery Mobile rendering."
        )
    )
    parser.add_argument(
        "--input",
        "-i",
        required=True,
        type=Path,
        help="CSV, XLSX, JSON, or text file containing scout names and QR URLs.",
    )
    parser.add_argument(
        "--sheet",
        default="Scouts Only",
        help='XLSX sheet name (default: "Scouts Only").',
    )
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        default=Path("scouting-output"),
        help="Output directory (default: scouting-output).",
    )
    parser.add_argument(
        "--headed",
        action="store_true",
        help="Show the Chromium window while running.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=45,
        help="Seconds to wait for the schedule response for each scout (default: 45).",
    )
    parser.add_argument(
        "--request-delay-ms",
        type=int,
        default=200,
        help="Delay between class-requirement requests (default: 200 ms).",
    )
    parser.add_argument(
        "--scout-delay-ms",
        type=int,
        default=500,
        help="Delay between scouts (default: 500 ms).",
    )
    parser.add_argument(
        "--allow-dead-assets",
        action="store_true",
        help="Do not block the two known mike.dev.* JavaScript URLs that currently time out.",
    )
    parser.add_argument(
        "--include-adults",
        action="store_true",
        help="Include Adult and Part-Time Adult rows when the input contains a Registrant Type column.",
    )
    parser.add_argument(
        "--no-html",
        action="store_true",
        help="Do not generate the self-contained sortable/filterable HTML report.",
    )
    parser.add_argument(
        "--html-name",
        default="report.html",
        help="HTML report filename inside the output directory (default: report.html).",
    )
    return parser.parse_args()


def norm_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def first_matching_key(headers: dict[str, int], candidates: Iterable[str]) -> int | None:
    for candidate in candidates:
        key = norm_header(candidate)
        if key in headers:
            return headers[key]
    return None


def load_inputs(path: Path, sheet_name: str, include_adults: bool) -> list[ScoutInput]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        scouts = load_xlsx(path, sheet_name)
    elif suffix == ".csv":
        scouts = load_csv(path)
    elif suffix == ".json":
        scouts = load_json(path)
    else:
        scouts = load_text(path)

    filtered: list[ScoutInput] = []
    seen: set[str] = set()
    for scout in scouts:
        scout.url = scout.url.strip()
        scout.name = scout.name.strip()
        scout.registrant_type = scout.registrant_type.strip()
        if not scout.url.startswith("https://scoutingevent.com/mobile/"):
            continue
        if not include_adults and scout.registrant_type.lower() in {"adult", "part-time adult"}:
            continue
        if scout.url in seen:
            continue
        seen.add(scout.url)
        filtered.append(scout)

    if not filtered:
        raise ValueError(f"No ScoutingEvent QR URLs found in {path}")
    return filtered


def load_xlsx(path: Path, sheet_name: str) -> list[ScoutInput]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    elif "Scouts Only" in workbook.sheetnames:
        sheet = workbook["Scouts Only"]
    else:
        sheet = workbook[workbook.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = {norm_header(value): index for index, value in enumerate(rows[0])}
    name_index = first_matching_key(headers, ["Attendee Name", "Name", "Scout Name"])
    url_index = first_matching_key(headers, ["QR Code Contents", "URL", "QR URL", "Link"])
    type_index = first_matching_key(headers, ["Registrant Type", "Type"])
    if url_index is None:
        raise ValueError(f"Could not find a URL column in XLSX sheet {sheet.title!r}")

    result: list[ScoutInput] = []
    for row in rows[1:]:
        url = str(row[url_index] or "")
        name = str(row[name_index] or "") if name_index is not None else ""
        registrant_type = str(row[type_index] or "") if type_index is not None else ""
        result.append(ScoutInput(name=name, url=url, registrant_type=registrant_type))
    return result


def load_csv(path: Path) -> list[ScoutInput]:
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        reader = csv.DictReader(stream)
        if reader.fieldnames:
            header_map = {norm_header(name): name for name in reader.fieldnames}
            url_key = next(
                (header_map[norm_header(name)] for name in ["QR Code Contents", "URL", "QR URL", "Link"] if norm_header(name) in header_map),
                None,
            )
            name_key = next(
                (header_map[norm_header(name)] for name in ["Attendee Name", "Name", "Scout Name"] if norm_header(name) in header_map),
                None,
            )
            type_key = next(
                (header_map[norm_header(name)] for name in ["Registrant Type", "Type"] if norm_header(name) in header_map),
                None,
            )
            if url_key:
                return [
                    ScoutInput(
                        name=(row.get(name_key, "") if name_key else "") or "",
                        url=(row.get(url_key, "") or ""),
                        registrant_type=(row.get(type_key, "") if type_key else "") or "",
                    )
                    for row in reader
                ]

    # Fall back to one URL per line if the CSV does not have recognizable headers.
    return load_text(path)


def load_json(path: Path) -> list[ScoutInput]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, dict):
        payload = payload.get("scouts") or payload.get("data") or [payload]
    result: list[ScoutInput] = []
    for item in payload:
        if isinstance(item, str):
            result.append(ScoutInput(name="", url=item))
        elif isinstance(item, dict):
            result.append(
                ScoutInput(
                    name=str(item.get("name") or item.get("attendee_name") or ""),
                    url=str(item.get("url") or item.get("qr_url") or item.get("qr_code_contents") or ""),
                    registrant_type=str(item.get("registrant_type") or item.get("type") or ""),
                )
            )
    return result


def load_text(path: Path) -> list[ScoutInput]:
    result: list[ScoutInput] = []
    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if "," in line and "https://" in line:
            name, url = line.split(",", 1)
            result.append(ScoutInput(name=name.strip(), url=url.strip()))
        else:
            result.append(ScoutInput(name="", url=line))
    return result


def form_values(post_data: str | None) -> dict[str, str]:
    parsed = parse_qs(post_data or "", keep_blank_values=True)
    return {key: values[-1] if values else "" for key, values in parsed.items()}


def is_schedule_response(response: Response) -> bool:
    if "ajax_mobile.php" not in response.url:
        return False
    values = form_values(response.request.post_data)
    return values.get("getUnitClass", "").lower() == "true"


def clean_html_description(value: str) -> str:
    value = re.sub(r"<br\s*/?>", "\n", value or "", flags=re.IGNORECASE)
    value = re.sub(r"<[^>]+>", "", value)
    value = html.unescape(value)
    value = re.sub(r"[ \t]+\n", "\n", value)
    value = re.sub(r"\n{3,}", "\n\n", value)
    return value.strip()


def truthy(value: Any) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y"}


def days_from_class_name(class_name: str) -> str:
    if " - " in class_name:
        return class_name.split(" - ", 1)[0].strip()
    return ""


async def discover_name(page: Page, fallback: str) -> str:
    if fallback:
        return fallback
    selectors = [
        "#scoutPage h1",
        "#scoutPage .ui-title",
        ".ui-page-active .ui-title",
        "h1.ui-title",
    ]
    for selector in selectors:
        locator = page.locator(selector)
        try:
            if await locator.count():
                text = (await locator.first.inner_text()).strip()
                if text and text.lower() not in {"schedule", "event", "menu"}:
                    return text
        except Exception:
            pass
    return "Unknown Scout"


async def navigate_and_trigger_schedule(page: Page, qr_url: str, timeout_seconds: int) -> Response:
    loop = asyncio.get_running_loop()
    future: asyncio.Future[Response] = loop.create_future()

    async def inspect_response(response: Response) -> None:
        if not future.done() and is_schedule_response(response):
            future.set_result(response)

    def on_response(response: Response) -> None:
        asyncio.create_task(inspect_response(response))

    page.on("response", on_response)
    try:
        # Attach the response listener before navigation so an eager page cannot race us.
        await page.goto(qr_url, wait_until="domcontentloaded", timeout=timeout_seconds * 1000)

        # The schedule sometimes loads automatically after the QR redirect.
        await page.wait_for_timeout(1200)
        if not future.done():
            selectors = [
                'a[href="#scoutPage"]',
                'a[href$="event.html#scoutPage"]',
                'a:has-text("Schedule")',
            ]
            clicked = False
            for selector in selectors:
                locator = page.locator(selector)
                try:
                    if await locator.count():
                        await locator.first.click(force=True, timeout=5000)
                        clicked = True
                        break
                except Exception:
                    continue
            if not clicked:
                # Invoke the existing handler in-page if the old UI is not considered clickable.
                await page.evaluate(
                    """() => {
                        const link = document.querySelector('a[href="#scoutPage"]') ||
                                     [...document.querySelectorAll('a')].find(a =>
                                         (a.textContent || '').trim().toLowerCase() === 'schedule');
                        if (link) link.click();
                    }"""
                )
        return await asyncio.wait_for(future, timeout=timeout_seconds)
    finally:
        page.remove_listener("response", on_response)


async def post_requirement(page: Page, attendee_id: str, class_p4_id: str) -> dict[str, Any]:
    result = await page.evaluate(
        """async ({url, attendeeId, classP4Id}) => {
            const body = new URLSearchParams({
                getClassRequirement: 'true',
                classP4ID: classP4Id,
                attendeeID: attendeeId,
                TimeStamp: new Date().toString()
            });
            const response = await fetch(url, {
                method: 'POST',
                credentials: 'same-origin',
                headers: {
                    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
                    'X-Requested-With': 'XMLHttpRequest'
                },
                body: body.toString()
            });
            return {
                ok: response.ok,
                status: response.status,
                text: await response.text()
            };
        }""",
        {"url": AJAX_URL, "attendeeId": attendee_id, "classP4Id": class_p4_id},
    )
    if not result["ok"]:
        raise RuntimeError(f"Requirement request returned HTTP {result['status']}")
    try:
        return json.loads(result["text"])
    except json.JSONDecodeError as exc:
        preview = result["text"][:300].replace("\n", " ")
        raise RuntimeError(f"Requirement response was not JSON: {preview}") from exc


async def process_scout(
    browser: Browser,
    scout: ScoutInput,
    args: argparse.Namespace,
    raw_dir: Path,
) -> ScoutResult:
    context: BrowserContext = await browser.new_context(ignore_https_errors=True)

    if not args.allow_dead_assets:
        async def block_dead_assets(route: Any) -> None:
            try:
                hostname = re.sub(r"^https?://", "", route.request.url).split("/", 1)[0].split(":", 1)[0]
                if hostname in DEAD_ASSET_HOSTS:
                    await route.abort()
                else:
                    await route.continue_()
            except Exception:
                await route.continue_()
        await context.route("**/*", block_dead_assets)

    page = await context.new_page()
    page.set_default_timeout(args.timeout * 1000)
    result = ScoutResult(name=scout.name or "Unknown Scout", qr_url=scout.url)

    page.on("pageerror", lambda error: print(f"    browser JS error: {error}", file=sys.stderr))

    try:
        schedule_response = await navigate_and_trigger_schedule(page, scout.url, args.timeout)
        request_values = form_values(schedule_response.request.post_data)
        attendee_id = request_values.get("attendeeID", "")
        if not attendee_id:
            raise RuntimeError("The schedule request did not contain attendeeID")

        result.attendee_id = attendee_id
        result.name = await discover_name(page, scout.name)

        schedule_payload = await schedule_response.json()
        raw_dir.mkdir(parents=True, exist_ok=True)
        slug = safe_slug(result.name)
        (raw_dir / f"{slug}-schedule.json").write_text(
            json.dumps(schedule_payload, indent=2, ensure_ascii=False), encoding="utf-8"
        )

        if int(schedule_payload.get("status", {}).get("status", 0)) != 1:
            raise RuntimeError(f"Schedule API returned failure: {schedule_payload.get('status')}")

        schedule_rows = schedule_payload.get("data") or []
        for item in schedule_rows:
            raw_class_name = str(item.get("CLASS_NAME") or "")
            class_record = ClassRecord(
                scout_name=result.name,
                attendee_id=attendee_id,
                class_number=str(item.get("CLASS_NBR") or ""),
                class_p4_id=str(item.get("classP4ID") or ""),
                class_name=str(item.get("CLASS_NAME_ONLY") or ""),
                days=days_from_class_name(raw_class_name),
                period_name=str(item.get("PERIOD_NAME") or ""),
                start_time=str(item.get("CLASS_STIME") or ""),
                start_time_24h=str(item.get("STIME") or ""),
                location=str(item.get("LOCATION_ROOM_NAME") or ""),
                has_requirements=truthy(item.get("HAS_REQ")),
                class_completed=truthy(item.get("CLASS_COMPLETED")),
                raw_class_name=raw_class_name,
            )
            result.classes.append(class_record)

            if not class_record.class_p4_id:
                continue

            requirement_payload: dict[str, Any] | None = None
            last_error: Exception | None = None
            for attempt in range(1, 4):
                try:
                    requirement_payload = await post_requirement(
                        page, attendee_id, class_record.class_p4_id
                    )
                    break
                except Exception as exc:
                    last_error = exc
                    await page.wait_for_timeout(500 * attempt)

            if requirement_payload is None:
                result.errors.append(
                    f"{class_record.class_name}: requirement request failed: {last_error}"
                )
                continue

            (raw_dir / f"{slug}-{safe_slug(class_record.class_number or class_record.class_p4_id)}-requirements.json").write_text(
                json.dumps(requirement_payload, indent=2, ensure_ascii=False), encoding="utf-8"
            )

            if int(requirement_payload.get("status", {}).get("status", 0)) != 1:
                result.errors.append(
                    f"{class_record.class_name}: requirement API returned failure: "
                    f"{requirement_payload.get('status')}"
                )
                continue

            for req in requirement_payload.get("data") or []:
                result.requirements.append(
                    RequirementRecord(
                        scout_name=result.name,
                        attendee_id=attendee_id,
                        class_number=class_record.class_number,
                        class_p4_id=class_record.class_p4_id,
                        merit_badge_name=str(req.get("MERIT_BADGE_NAME") or ""),
                        requirement_id=str(req.get("MDM_LKP_BADGE_REQ_ID") or ""),
                        requirement_type_id=str(req.get("MDM_LKP_BADGE_REQ_TYPE_ID") or ""),
                        requirement_number=str(req.get("REQ_NBR_WEB_DISPLAY") or req.get("REQ_NBR") or ""),
                        description=clean_html_description(str(req.get("REQ_DESCR") or "")),
                        completed_requirement=truthy(req.get("COMPLETED_REQ_FLAG")),
                        completed_class=truthy(req.get("COMPLETED_CLASS_FLAG")),
                        requirement_on_event=truthy(req.get("REQ_ON_EVENT")),
                        present_days=str(req.get("PRESENT_DAYS") or "").strip(),
                    )
                )

            if args.request_delay_ms > 0:
                await page.wait_for_timeout(args.request_delay_ms)

    except Exception as exc:
        result.errors.append(str(exc))
    finally:
        await context.close()

    return result


def safe_slug(value: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "-", value.strip()).strip("-").lower()
    return slug or "unknown-scout"


def write_csv(path: Path, records: list[Any], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for record in records:
            writer.writerow(asdict(record))




CHOICE_COUNT_WORDS = {
    "one": 1,
    "two": 2,
    "three": 3,
    "four": 4,
    "five": 5,
    "six": 6,
    "seven": 7,
    "eight": 8,
    "nine": 9,
    "ten": 10,
    "eleven": 11,
    "twelve": 12,
}


def requirement_path(number: str, current_top_level: str | None) -> tuple[tuple[str, ...], str | None]:
    """Return a comparable requirement path and the current top-level number.

    ScoutingEvent occasionally emits malformed display numbers such as ``#(c)``
    where the leading top-level number is omitted. In that case, retain the most
    recently observed numeric top-level requirement so nested rows still group
    correctly in the HTML report.
    """
    text = str(number or "").strip()
    tokens = [token.lower() for token in re.findall(r"[A-Za-z]+|\d+", text)]
    if not tokens:
        return (), current_top_level

    if tokens[0].isdigit():
        current_top_level = tokens[0]
        return tuple(tokens), current_top_level

    if current_top_level:
        return (current_top_level, *tokens), current_top_level

    return tuple(tokens), current_top_level


def choice_requirement_count(description: str) -> int | None:
    """Extract X from wording such as 'Do TWO of the following'."""
    text = re.sub(r"\s+", " ", str(description or "")).strip().lower()
    words = "|".join(CHOICE_COUNT_WORDS)
    match = re.search(
        rf"\b(?:do|complete|choose|select|perform)\s+(?:any\s+)?(?P<count>\d+|{words})\s+of\s+the\s+following\b",
        text,
        flags=re.IGNORECASE,
    )
    if not match:
        return None
    value = match.group("count").lower()
    return int(value) if value.isdigit() else CHOICE_COUNT_WORDS.get(value)


def annotate_requirement_statuses(requirements: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Add inferred display status fields used only by the HTML report.

    The API frequently leaves parent rows incomplete even when every child row is
    complete. Parent status is therefore calculated recursively. Choice parents
    such as 'Do ONE of the following' are marked as completed-but-review when the
    minimum number of alternatives appears complete.
    """
    annotated = [dict(item) for item in requirements]
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for item in annotated:
        grouped.setdefault((str(item.get("scout_name") or ""), str(item.get("class_p4_id") or "")), []).append(item)

    for group in grouped.values():
        current_top_level: str | None = None
        paths: list[tuple[str, ...]] = []
        for item in group:
            path, current_top_level = requirement_path(str(item.get("requirement_number") or ""), current_top_level)
            paths.append(path)
            item["requirement_path"] = list(path)

        children: list[list[int]] = [[] for _ in group]
        prior_path_indexes: dict[tuple[str, ...], int] = {}
        header_stack: list[tuple[int, int]] = []

        for index, (item, path) in enumerate(zip(group, paths)):
            parent_index: int | None = None
            if len(path) > 1:
                for depth in range(len(path) - 1, 0, -1):
                    candidate = path[:depth]
                    if candidate in prior_path_indexes:
                        parent_index = prior_path_indexes[candidate]
                        break

            # Fallback for malformed numbering: use the nearest prior header with
            # a shallower inferred path. This preserves ordering without making
            # assumptions across unrelated top-level requirements.
            if parent_index is None and len(path) > 1:
                for candidate_index, candidate_depth in reversed(header_stack):
                    if candidate_depth < len(path):
                        parent_index = candidate_index
                        break

            if parent_index is not None:
                children[parent_index].append(index)

            if path:
                prior_path_indexes[path] = index
            if str(item.get("requirement_type_id") or "") == "3":
                while header_stack and header_stack[-1][1] >= len(path):
                    header_stack.pop()
                header_stack.append((index, len(path)))

        cache: dict[int, tuple[str, str]] = {}

        def calculate(index: int) -> tuple[str, str]:
            if index in cache:
                return cache[index]

            item = group[index]
            direct_children = children[index]
            if bool(item.get("completed_requirement")):
                result = ("complete", "Marked complete by ScoutingEvent.")
            elif direct_children:
                child_results = [calculate(child_index) for child_index in direct_children]
                satisfied = [status in {"complete", "complete_check"} for status, _ in child_results]
                satisfied_count = sum(satisfied)
                total_count = len(direct_children)
                choice_count = choice_requirement_count(str(item.get("description") or ""))

                item["completed_subrequirements"] = satisfied_count
                item["total_subrequirements"] = total_count
                item["choice_required_count"] = choice_count

                if total_count and satisfied_count == total_count:
                    if any(status == "complete_check" for status, _ in child_results):
                        result = (
                            "complete_check",
                            f"All {total_count} subrequirements appear satisfied, but at least one was inferred from a choice requirement.",
                        )
                    else:
                        result = ("complete", f"All {total_count} subrequirements are complete.")
                elif choice_count is not None and satisfied_count >= choice_count:
                    result = (
                        "complete_check",
                        f"{satisfied_count} of {total_count} subrequirements appear complete; this meets the stated choice of {choice_count}, but should be verified.",
                    )
                else:
                    result = (
                        "incomplete",
                        f"{satisfied_count} of {total_count} subrequirements are complete.",
                    )
            else:
                result = ("incomplete", "Not marked complete by ScoutingEvent.")

            cache[index] = result
            return result

        for index, item in enumerate(group):
            status, reason = calculate(index)
            item["calculated_status"] = status
            item["calculated_status_reason"] = reason

    return annotated

def json_for_html(value: Any) -> str:
    """Serialize JSON safely for embedding inside an HTML script element."""
    return (
        json.dumps(value, ensure_ascii=False, separators=(",", ":"))
        .replace("&", "\\u0026")
        .replace("<", "\\u003c")
        .replace(">", "\\u003e")
        .replace("\u2028", "\\u2028")
        .replace("\u2029", "\\u2029")
    )


def write_html_report(
    path: Path,
    results: list[ScoutResult],
    generated_at_local: str,
) -> None:
    classes = [asdict(item) for result in results for item in result.classes]
    requirements = annotate_requirement_statuses(
        [asdict(item) for result in results for item in result.requirements]
    )
    errors = [
        {"scout_name": result.name, "attendee_id": result.attendee_id, "error": error}
        for result in results
        for error in result.errors
    ]
    scouts = [
        {
            "name": result.name,
            "attendee_id": result.attendee_id,
            "class_count": len(result.classes),
            "requirement_count": len(result.requirements),
            "error_count": len(result.errors),
        }
        for result in results
    ]
    report_data = {
        "generated_at_local": generated_at_local,
        "scouts": scouts,
        "classes": classes,
        "requirements": requirements,
        "errors": errors,
    }

    template = r'''<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>ScoutingEvent Schedule Report</title>
<style>
:root {
  color-scheme: light dark;
  --bg: #f4f6f8;
  --panel: #ffffff;
  --panel-alt: #eef3f6;
  --text: #17212b;
  --muted: #5e6b78;
  --border: #d3dbe3;
  --accent: #245b78;
  --accent-2: #dbeaf2;
  --success: #176b3a;
  --success-bg: #dff2e7;
  --warn: #8a4b08;
  --warn-bg: #fff0d8;
  --danger: #9b1c1c;
  --danger-bg: #fde5e5;
  --shadow: 0 2px 10px rgba(18, 38, 54, 0.08);
}
@media (prefers-color-scheme: dark) {
  :root {
    --bg: #111820;
    --panel: #18222d;
    --panel-alt: #202d39;
    --text: #ecf2f7;
    --muted: #aab7c4;
    --border: #334353;
    --accent: #70b6db;
    --accent-2: #203d4d;
    --success: #8fe0ad;
    --success-bg: #173b26;
    --warn: #ffd08a;
    --warn-bg: #4a3214;
    --danger: #ffaaaa;
    --danger-bg: #4b2222;
    --shadow: none;
  }
}
* { box-sizing: border-box; }
body {
  margin: 0;
  background: var(--bg);
  color: var(--text);
  font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
  line-height: 1.45;
}
header {
  background: linear-gradient(135deg, #173f57, #28739a);
  color: white;
  padding: 1.5rem clamp(1rem, 4vw, 3rem);
}
header h1 { margin: 0; font-size: clamp(1.55rem, 3vw, 2.35rem); }
header p { margin: .35rem 0 0; opacity: .9; }
main { max-width: 1500px; margin: 0 auto; padding: 1rem; }
.summary-grid {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(145px, 1fr));
  gap: .75rem;
  margin-bottom: 1rem;
}
.summary-card, .panel {
  background: var(--panel);
  border: 1px solid var(--border);
  border-radius: 10px;
  box-shadow: var(--shadow);
}
.summary-card { padding: .9rem 1rem; }
.summary-card .value { font-size: 1.6rem; font-weight: 750; }
.summary-card .label { color: var(--muted); font-size: .85rem; }
.panel { padding: 1rem; margin-bottom: 1rem; }
.controls {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(175px, 1fr));
  gap: .75rem;
  align-items: end;
}
.control label { display: block; font-size: .78rem; font-weight: 700; color: var(--muted); margin-bottom: .25rem; }
input, select, button {
  width: 100%;
  border: 1px solid var(--border);
  border-radius: 7px;
  background: var(--panel);
  color: var(--text);
  padding: .62rem .7rem;
  font: inherit;
}
.checkbox-control { align-self: end; }
.checkbox-control[hidden] { display: none; }
.checkbox-label {
  display: flex !important;
  align-items: center;
  gap: .55rem;
  min-height: 45px;
  margin: 0 !important;
  padding: .55rem .7rem;
  border: 1px solid var(--border);
  border-radius: 7px;
  background: var(--panel-alt);
  color: var(--text) !important;
  cursor: pointer;
}
.checkbox-label input {
  width: auto;
  margin: 0;
  padding: 0;
  accent-color: var(--accent);
}
.checkbox-label span { line-height: 1.25; }
button { cursor: pointer; font-weight: 700; }
button:hover { border-color: var(--accent); }
.tabs { display: flex; gap: .4rem; flex-wrap: wrap; margin-bottom: 1rem; }
.tabs button { width: auto; padding-inline: 1rem; }
.tabs button.active { background: var(--accent); color: white; border-color: var(--accent); }
.view[hidden] { display: none; }
.result-count { color: var(--muted); margin: 0 0 .75rem; }
.scout-card {
  background: var(--panel);
  border: 1px solid var(--border);
  border-radius: 10px;
  margin-bottom: .9rem;
  overflow: hidden;
}
.scout-card > h2 {
  margin: 0;
  padding: .85rem 1rem;
  background: var(--panel-alt);
  border-bottom: 1px solid var(--border);
  font-size: 1.1rem;
  display: flex;
  justify-content: space-between;
  gap: 1rem;
}
.scout-card h2 small { color: var(--muted); font-weight: 500; }
.class-list { display: grid; }
.class-row {
  display: grid;
  grid-template-columns: 105px 85px minmax(260px, 2fr) 130px minmax(170px, 1fr) auto;
  gap: .7rem;
  align-items: center;
  padding: .7rem 1rem;
  border-bottom: 1px solid var(--border);
}
.class-row:last-child { border-bottom: 0; }
.class-row:hover { background: var(--panel-alt); }
.class-name { font-weight: 720; }
.meta { color: var(--muted); font-size: .86rem; }
.badge {
  display: inline-block;
  padding: .16rem .5rem;
  border-radius: 999px;
  font-size: .76rem;
  font-weight: 750;
  white-space: nowrap;
}
.badge.success { color: var(--success); background: var(--success-bg); }
.badge.warn { color: var(--warn); background: var(--warn-bg); }
.badge.danger { color: var(--danger); background: var(--danger-bg); }
.badge.check { color: var(--warn); background: var(--warn-bg); border: 1px solid var(--warn); }
details.requirements { grid-column: 1 / -1; }
details.requirements summary { cursor: pointer; color: var(--accent); font-weight: 700; }
.requirement-list { margin: .6rem 0 0; padding: 0; list-style: none; }
.requirement-list li { border-top: 1px solid var(--border); padding: .5rem 0; }
.requirement-number { font-weight: 750; margin-right: .35rem; }
.table-wrap { overflow: auto; max-height: 72vh; border: 1px solid var(--border); border-radius: 8px; }
table { border-collapse: separate; border-spacing: 0; width: 100%; min-width: 920px; background: var(--panel); }
th, td { text-align: left; padding: .62rem .7rem; border-bottom: 1px solid var(--border); vertical-align: top; }
th {
  position: sticky;
  top: 0;
  z-index: 2;
  background: var(--panel-alt);
  font-size: .8rem;
  text-transform: uppercase;
  letter-spacing: .03em;
  cursor: pointer;
  user-select: none;
}
th.sorted::after { content: "  ▲"; }
th.sorted.desc::after { content: "  ▼"; }
tr:hover td { background: var(--panel-alt); }
.empty { padding: 2rem; text-align: center; color: var(--muted); }
.error-box { border-left: 4px solid var(--danger); background: var(--danger-bg); padding: .75rem 1rem; margin-top: 1rem; }
@media (max-width: 900px) {
  .class-row { grid-template-columns: 90px 70px 1fr; }
  .class-row > :nth-child(4), .class-row > :nth-child(5), .class-row > :nth-child(6) { grid-column: 3; }
}
@media print {
  body { background: white; color: black; }
  header { background: white; color: black; padding: 0 0 1rem; }
  .controls, .tabs, #classes-view, #requirements-view { display: none !important; }
  main { max-width: none; padding: 0; }
  .panel, .scout-card { box-shadow: none; break-inside: avoid; }
}
</style>
</head>
<body>
<header>
  <h1>ScoutingEvent Schedule Report</h1>
  <p id="generated-at"></p>
</header>
<main>
  <section class="summary-grid" id="summary"></section>

  <section class="panel">
    <div class="controls">
      <div class="control">
        <label for="search">Search</label>
        <input id="search" type="search" placeholder="Scout, class, number, location…">
      </div>
      <div class="control">
        <label for="scout-filter">Scout</label>
        <select id="scout-filter"><option value="">All scouts</option></select>
      </div>
      <div class="control">
        <label for="class-filter">Class</label>
        <select id="class-filter"><option value="">All classes</option></select>
      </div>
      <div class="control">
        <label for="block-filter">Block</label>
        <select id="block-filter"><option value="">All blocks</option></select>
      </div>
      <div class="control">
        <label for="day-filter">Day</label>
        <select id="day-filter"><option value="">All days</option></select>
      </div>
      <div class="control">
        <label for="location-filter">Location</label>
        <select id="location-filter"><option value="">All locations</option></select>
      </div>
      <div class="control">
        <label for="completion-filter">Completion</label>
        <select id="completion-filter">
          <option value="">All</option>
          <option value="complete">Completed classes</option>
          <option value="incomplete">Incomplete classes</option>
        </select>
      </div>
      <div class="control checkbox-control" id="completed-requirements-control" hidden>
        <label class="checkbox-label" for="show-completed-requirements">
          <input id="show-completed-requirements" type="checkbox">
          <span>Show completed requirements in incomplete classes</span>
        </label>
      </div>
      <div class="control">
        <label for="sort-select">Sort scout schedules</label>
        <select id="sort-select">
          <option value="time">Start time</option>
          <option value="class">Class name</option>
          <option value="block">Block</option>
          <option value="location">Location</option>
        </select>
      </div>
      <div class="control"><label>&nbsp;</label><button id="reset">Reset filters</button></div>
    </div>
  </section>

  <nav class="tabs" aria-label="Report views">
    <button class="active" data-view="scouts-view">By Scout</button>
    <button data-view="classes-view">All Classes</button>
    <button data-view="requirements-view">Requirements</button>
  </nav>

  <section id="scouts-view" class="view">
    <p class="result-count" id="scout-count"></p>
    <div id="scout-results"></div>
  </section>

  <section id="classes-view" class="view" hidden>
    <p class="result-count" id="class-count"></p>
    <div class="table-wrap">
      <table id="classes-table">
        <thead><tr>
          <th data-key="scout_name">Scout</th>
          <th data-key="start_time_24h">Time</th>
          <th data-key="days">Days</th>
          <th data-key="period_name">Block</th>
          <th data-key="class_name">Class</th>
          <th data-key="class_number">Class #</th>
          <th data-key="location">Location</th>
          <th data-key="class_completed">Status</th>
        </tr></thead>
        <tbody></tbody>
      </table>
    </div>
  </section>

  <section id="requirements-view" class="view" hidden>
    <p class="result-count" id="requirement-count"></p>
    <div class="table-wrap">
      <table id="requirements-table">
        <thead><tr>
          <th data-key="scout_name">Scout</th>
          <th data-key="merit_badge_name">Class</th>
          <th data-key="class_number">Class #</th>
          <th data-key="requirement_number">Requirement</th>
          <th data-key="description">Description</th>
          <th data-key="present_days">Days</th>
          <th data-key="completed_requirement">Complete</th>
          <th data-key="requirement_on_event">At Event</th>
        </tr></thead>
        <tbody></tbody>
      </table>
    </div>
  </section>

  <section id="errors"></section>
</main>
<script id="report-data" type="application/json">__REPORT_DATA__</script>
<script>
(() => {
  "use strict";
  const data = JSON.parse(document.getElementById("report-data").textContent);
  const classes = data.classes || [];
  const requirements = data.requirements || [];
  const scouts = data.scouts || [];
  const dayOrder = ["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"];
  const dayLabels = {Mo:"Monday", Tu:"Tuesday", We:"Wednesday", Th:"Thursday", Fr:"Friday", Sa:"Saturday", Su:"Sunday"};
  const state = { classSort: {key:"scout_name", desc:false}, reqSort: {key:"scout_name", desc:false} };

  const $ = id => document.getElementById(id);
  const escapeHtml = value => String(value ?? "")
    .replaceAll("&", "&amp;").replaceAll("<", "&lt;").replaceAll(">", "&gt;")
    .replaceAll('"', "&quot;").replaceAll("'", "&#039;");
  const boolBadge = (value, yes="Complete", no="Incomplete") => value
    ? `<span class="badge success">${escapeHtml(yes)}</span>`
    : `<span class="badge warn">${escapeHtml(no)}</span>`;
  const requirementBadge = item => {
    const reason = escapeHtml(item.calculated_status_reason || "");
    if (item.calculated_status === "complete") {
      return `<span class="badge success" title="${reason}">Completed</span>`;
    }
    if (item.calculated_status === "complete_check") {
      return `<span class="badge check" title="${reason}">Completed, but double check</span>`;
    }
    return `<span class="badge warn" title="${reason}">Not complete</span>`;
  };
  const requirementIsComplete = item => ["complete", "complete_check"].includes(item.calculated_status);
  const tokensForDays = value => dayOrder.filter(token => String(value || "").includes(token));
  const normalize = value => String(value ?? "").trim().toLocaleLowerCase();
  const uniqueSorted = values => [...new Set(values.filter(Boolean))].sort((a,b) => a.localeCompare(b, undefined, {numeric:true}));
  const reqKey = item => `${item.scout_name}\u0000${item.class_p4_id}`;
  const requirementsByClass = new Map();
  for (const req of requirements) {
    const key = reqKey(req);
    if (!requirementsByClass.has(key)) requirementsByClass.set(key, []);
    requirementsByClass.get(key).push(req);
  }

  $("generated-at").textContent = `Generated ${data.generated_at_local || ""}`;
  $("summary").innerHTML = [
    [scouts.length, "Scouts"],
    [classes.length, "Scheduled classes"],
    [requirements.length, "Requirements"],
    [data.errors?.length || 0, "Errors"]
  ].map(([value,label]) => `<div class="summary-card"><div class="value">${value}</div><div class="label">${label}</div></div>`).join("");

  function fillSelect(id, values, labels = null) {
    const select = $(id);
    for (const value of values) {
      const option = document.createElement("option");
      option.value = value;
      option.textContent = labels?.[value] || value;
      select.appendChild(option);
    }
  }
  fillSelect("scout-filter", uniqueSorted(scouts.map(s => s.name)));
  fillSelect("class-filter", uniqueSorted(classes.map(c => c.class_name)));
  fillSelect("block-filter", uniqueSorted(classes.map(c => c.period_name)));
  fillSelect("day-filter", dayOrder.filter(day => classes.some(c => tokensForDays(c.days).includes(day))), dayLabels);
  fillSelect("location-filter", uniqueSorted(classes.map(c => c.location)));

  function filters() {
    return {
      search: normalize($("search").value), scout: $("scout-filter").value,
      className: $("class-filter").value, block: $("block-filter").value,
      day: $("day-filter").value, location: $("location-filter").value,
      completion: $("completion-filter").value,
      showCompletedRequirements: $("show-completed-requirements").checked
    };
  }

  function classMatches(item, f = filters()) {
    const haystack = normalize([item.scout_name, item.class_name, item.class_number, item.location, item.period_name, item.days].join(" "));
    return (!f.search || haystack.includes(f.search))
      && (!f.scout || item.scout_name === f.scout)
      && (!f.className || item.class_name === f.className)
      && (!f.block || item.period_name === f.block)
      && (!f.day || tokensForDays(item.days).includes(f.day))
      && (!f.location || item.location === f.location)
      && (!f.completion || (f.completion === "complete") === Boolean(item.class_completed));
  }

  function requirementMatches(item, f = filters()) {
    const linkedClass = classes.find(c => c.scout_name === item.scout_name && c.class_p4_id === item.class_p4_id);
    const haystack = normalize([item.scout_name, item.merit_badge_name, item.class_number, item.requirement_number, item.description, item.present_days].join(" "));
    return (!f.search || haystack.includes(f.search))
      && (!f.scout || item.scout_name === f.scout)
      && (!f.className || item.merit_badge_name === f.className || linkedClass?.class_name === f.className)
      && (!f.block || linkedClass?.period_name === f.block)
      && (!f.day || tokensForDays(item.present_days).includes(f.day) || tokensForDays(linkedClass?.days).includes(f.day))
      && (!f.location || linkedClass?.location === f.location)
      && (!f.completion || (f.completion === "complete") === Boolean(item.completed_class))
      && (!(f.completion === "incomplete" && !f.showCompletedRequirements) || !requirementIsComplete(item));
  }

  function compareValues(a, b, key, desc) {
    let av = a[key], bv = b[key];
    if (typeof av === "boolean" || typeof bv === "boolean") { av = Number(Boolean(av)); bv = Number(Boolean(bv)); }
    const result = String(av ?? "").localeCompare(String(bv ?? ""), undefined, {numeric:true, sensitivity:"base"});
    return desc ? -result : result;
  }

  function renderScouts() {
    const f = filters();
    const sortKey = $("sort-select").value;
    const sortMap = {time:"start_time_24h", class:"class_name", block:"period_name", location:"location"};
    const grouped = new Map();
    for (const item of classes.filter(c => classMatches(c, f))) {
      if (!grouped.has(item.scout_name)) grouped.set(item.scout_name, []);
      grouped.get(item.scout_name).push(item);
    }
    const scoutNames = [...grouped.keys()].sort((a,b) => a.localeCompare(b));
    $("scout-count").textContent = `${scoutNames.length} scout(s), ${[...grouped.values()].reduce((n,v)=>n+v.length,0)} matching class(es)`;
    if (!scoutNames.length) {
      $("scout-results").innerHTML = '<div class="panel empty">No classes match the current filters.</div>';
      return;
    }
    $("scout-results").innerHTML = scoutNames.map(name => {
      const rows = grouped.get(name).slice().sort((a,b) => compareValues(a,b,sortMap[sortKey],false) || compareValues(a,b,"class_name",false));
      const content = rows.map(item => {
        const allReqs = requirementsByClass.get(reqKey(item)) || [];
        const hideCompleted = f.completion === "incomplete" && !f.showCompletedRequirements;
        const reqs = hideCompleted ? allReqs.filter(req => !requirementIsComplete(req)) : allReqs;
        const hiddenCount = allReqs.length - reqs.length;
        let reqHtml = "";
        if (allReqs.length) {
          const hiddenText = hiddenCount ? `; ${hiddenCount} completed hidden` : "";
          const reqBody = reqs.length
            ? `<ul class="requirement-list">${reqs.map(req => `<li><span class="requirement-number">${escapeHtml(req.requirement_number)}</span>${escapeHtml(req.description)} ${requirementBadge(req)}</li>`).join("")}</ul>`
            : `<p class="meta">All ${allReqs.length} requirement row(s) are completed and currently hidden.</p>`;
          reqHtml = `<details class="requirements"><summary>${reqs.length} visible requirement row(s)${hiddenText}</summary>${reqBody}</details>`;
        }
        return `<div class="class-row">
          <div><strong>${escapeHtml(item.start_time)}</strong></div>
          <div>${escapeHtml(item.days)}</div>
          <div><div class="class-name">${escapeHtml(item.class_name)}</div><div class="meta">${escapeHtml(item.class_number)}</div></div>
          <div>${escapeHtml(item.period_name)}</div>
          <div>${escapeHtml(item.location)}</div>
          <div>${boolBadge(item.class_completed)}</div>
          ${reqHtml}
        </div>`;
      }).join("");
      return `<article class="scout-card"><h2><span>${escapeHtml(name)}</span><small>${rows.length} class(es)</small></h2><div class="class-list">${content}</div></article>`;
    }).join("");
  }

  function renderClasses() {
    const rows = classes.filter(item => classMatches(item)).slice().sort((a,b) => compareValues(a,b,state.classSort.key,state.classSort.desc));
    $("class-count").textContent = `${rows.length} matching class row(s)`;
    $("classes-table").querySelector("tbody").innerHTML = rows.map(item => `<tr>
      <td>${escapeHtml(item.scout_name)}</td><td>${escapeHtml(item.start_time)}</td><td>${escapeHtml(item.days)}</td>
      <td>${escapeHtml(item.period_name)}</td><td><strong>${escapeHtml(item.class_name)}</strong></td>
      <td>${escapeHtml(item.class_number)}</td><td>${escapeHtml(item.location)}</td><td>${boolBadge(item.class_completed)}</td>
    </tr>`).join("") || '<tr><td colspan="8" class="empty">No classes match the current filters.</td></tr>';
    updateSortHeaders("classes-table", state.classSort);
  }

  function renderRequirements() {
    const rows = requirements.filter(item => requirementMatches(item)).slice().sort((a,b) => compareValues(a,b,state.reqSort.key,state.reqSort.desc));
    $("requirement-count").textContent = `${rows.length} matching requirement row(s)`;
    $("requirements-table").querySelector("tbody").innerHTML = rows.map(item => `<tr>
      <td>${escapeHtml(item.scout_name)}</td><td>${escapeHtml(item.merit_badge_name)}</td><td>${escapeHtml(item.class_number)}</td>
      <td><strong>${escapeHtml(item.requirement_number)}</strong></td><td>${escapeHtml(item.description).replaceAll("\n", "<br>")}</td>
      <td>${escapeHtml(item.present_days)}</td><td>${requirementBadge(item)}</td>
      <td>${boolBadge(item.requirement_on_event, "Yes", "No")}</td>
    </tr>`).join("") || '<tr><td colspan="8" class="empty">No requirements match the current filters.</td></tr>';
    updateSortHeaders("requirements-table", state.reqSort);
  }

  function updateSortHeaders(tableId, sort) {
    for (const th of $(tableId).querySelectorAll("th[data-key]")) {
      th.classList.toggle("sorted", th.dataset.key === sort.key);
      th.classList.toggle("desc", th.dataset.key === sort.key && sort.desc);
    }
  }

  function syncCompletedRequirementsControl() {
    $("completed-requirements-control").hidden = $("completion-filter").value !== "incomplete";
  }

  function renderAll() {
    syncCompletedRequirementsControl();
    renderScouts();
    renderClasses();
    renderRequirements();
  }
  for (const id of ["search","scout-filter","class-filter","block-filter","day-filter","location-filter","completion-filter","sort-select","show-completed-requirements"]) {
    $(id).addEventListener(id === "search" ? "input" : "change", renderAll);
  }
  $("reset").addEventListener("click", () => {
    for (const id of ["search","scout-filter","class-filter","block-filter","day-filter","location-filter","completion-filter"]) $(id).value = "";
    $("show-completed-requirements").checked = false;
    $("sort-select").value = "time";
    renderAll();
  });
  for (const button of document.querySelectorAll(".tabs button")) {
    button.addEventListener("click", () => {
      document.querySelectorAll(".tabs button").forEach(b => b.classList.toggle("active", b === button));
      document.querySelectorAll(".view").forEach(view => view.hidden = view.id !== button.dataset.view);
    });
  }
  for (const [tableId, sortState] of [["classes-table", state.classSort], ["requirements-table", state.reqSort]]) {
    for (const th of $(tableId).querySelectorAll("th[data-key]")) {
      th.addEventListener("click", () => {
        if (sortState.key === th.dataset.key) sortState.desc = !sortState.desc;
        else { sortState.key = th.dataset.key; sortState.desc = false; }
        tableId === "classes-table" ? renderClasses() : renderRequirements();
      });
    }
  }
  if (data.errors?.length) {
    $("errors").innerHTML = `<div class="error-box"><strong>${data.errors.length} extraction error(s)</strong><ul>${data.errors.map(e => `<li>${escapeHtml(e.scout_name)}: ${escapeHtml(e.error)}</li>`).join("")}</ul></div>`;
  }
  renderAll();
})();
</script>
</body>
</html>
'''
    path.write_text(
        template.replace("__REPORT_DATA__", json_for_html(report_data)),
        encoding="utf-8",
    )


def write_outputs(
    output_dir: Path,
    results: list[ScoutResult],
    *,
    generate_html: bool = True,
    html_name: str = "report.html",
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    classes = [item for result in results for item in result.classes]
    requirements = [item for result in results for item in result.requirements]
    generated_at_local = time.strftime("%Y-%m-%d %H:%M:%S %Z")

    nested = []
    for result in results:
        nested.append(
            {
                "name": result.name,
                "qr_url": result.qr_url,
                "attendee_id": result.attendee_id,
                "classes": [asdict(item) for item in result.classes],
                "requirements": [asdict(item) for item in result.requirements],
                "errors": result.errors,
            }
        )
    (output_dir / "scouts.json").write_text(
        json.dumps(nested, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    write_csv(
        output_dir / "classes.csv",
        classes,
        [field.name for field in ClassRecord.__dataclass_fields__.values()],
    )
    write_csv(
        output_dir / "requirements.csv",
        requirements,
        [field.name for field in RequirementRecord.__dataclass_fields__.values()],
    )

    error_rows = [
        {"scout_name": result.name, "attendee_id": result.attendee_id, "error": error}
        for result in results
        for error in result.errors
    ]
    with (output_dir / "errors.csv").open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=["scout_name", "attendee_id", "error"])
        writer.writeheader()
        writer.writerows(error_rows)

    safe_html_name = Path(html_name).name or "report.html"
    if not safe_html_name.lower().endswith(".html"):
        safe_html_name += ".html"
    if generate_html:
        write_html_report(output_dir / safe_html_name, results, generated_at_local)

    summary = {
        "scouts_processed": len(results),
        "scouts_with_errors": sum(1 for result in results if result.errors),
        "classes": len(classes),
        "requirements": len(requirements),
        "generated_at_local": generated_at_local,
        "html_report": safe_html_name if generate_html else None,
    }
    (output_dir / "summary.json").write_text(
        json.dumps(summary, indent=2), encoding="utf-8"
    )


async def async_main(args: argparse.Namespace) -> int:
    scouts = load_inputs(args.input, args.sheet, args.include_adults)
    args.output.mkdir(parents=True, exist_ok=True)
    raw_dir = args.output / "raw"

    print(f"Loaded {len(scouts)} attendee QR link(s).")
    results: list[ScoutResult] = []

    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(headless=not args.headed)
        try:
            for index, scout in enumerate(scouts, start=1):
                label = scout.name or scout.url
                print(f"[{index}/{len(scouts)}] {label}")
                result = await process_scout(browser, scout, args, raw_dir)
                results.append(result)
                print(
                    f"    attendeeID={result.attendee_id or 'unknown'}, "
                    f"classes={len(result.classes)}, requirements={len(result.requirements)}, "
                    f"errors={len(result.errors)}"
                )
                if args.scout_delay_ms > 0 and index < len(scouts):
                    await asyncio.sleep(args.scout_delay_ms / 1000)
        finally:
            await browser.close()

    write_outputs(
        args.output,
        results,
        generate_html=not args.no_html,
        html_name=args.html_name,
    )
    print(f"Results written to: {args.output.resolve()}")
    return 1 if any(result.errors for result in results) else 0


def main() -> int:
    args = parse_args()
    try:
        return asyncio.run(async_main(args))
    except KeyboardInterrupt:
        print("Interrupted.", file=sys.stderr)
        return 130
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
